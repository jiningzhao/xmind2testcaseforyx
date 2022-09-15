#!/usr/bin/env python
# _*_ coding:utf-8 _*_
import sys
from io import StringIO
import pandas as pd
from openpyxl import load_workbook

import xlwt

sys.path.append("/Users/zhaojining/PycharmProjects/pythonProject")
import csv
import logging
import os
from xmind2testcaseforyx.utils import get_xmind_testcase_list, get_absolute_path

"""
Convert XMind fie to Zentao testcase csv file 

Zentao official document about import CSV testcase file: https://www.zentao.net/book/zentaopmshelp/243.mhtml 
"""


def xmind_to_yaxin_xls_file(xmind_file):
    """Convert XMind file to a zentao csv file"""
    xmind_file = get_absolute_path(xmind_file)
    logging.info('Start converting XMind file(%s) to zentao file...', xmind_file)
    testcases = get_xmind_testcase_list(xmind_file)

    # fileheader = ["所属模块", "用例标题", "前置条件", "步骤", "预期", "关键词", "优先级", "用例类型", "适用阶段"]
    fileheader = ["用例名称", "需求号", "测试范围", "模块名称", "优先级", "为核心用例", "已自动化", "为回归用例",
                  "前提条件", "操作步骤", "期望结果", "执行结果"]
    yaxin_testcase_rows = [fileheader]
    for testcase in testcases:
        rows = gen_a_testcase_row(testcase)
        for row in rows:
            yaxin_testcase_rows.append(row)

    yaxin_csv_file = xmind_file[:-6] + '.csv'
    yaxin_xls_file = xmind_file[:-6] + '.xlsx'

    if os.path.exists(yaxin_csv_file):
        os.remove(yaxin_csv_file)
        # logging.info('The zentao csv file already exists, return it directly: %s', yaxin_file)
        # return yaxin_file

    with open(yaxin_csv_file, 'w', encoding='gbk') as f:
        writer = csv.writer(f)
        writer.writerows(yaxin_testcase_rows)
        logging.info('Convert XMind file(%s) to a yaxin xls file(%s) successfully!', xmind_file, yaxin_csv_file)

    return csv_to_xls(yaxin_csv_file, yaxin_xls_file)


def gen_a_testcase_row(testcase_dict):
    # 所属模块
    try:
        case_module, requirement_about_case = gen_case_module(testcase_dict['suite']).split("#")
    except Exception:
        case_module, requirement_about_case = gen_case_module(testcase_dict['suite']), ""
    # 用例标题
    case_title = testcase_dict['name']
    # 前置条件
    case_precontion = testcase_dict['preconditions']
    # 步骤+预期
    case_step, case_expected_result = gen_case_step_and_expected_result(testcase_dict['steps'])
    case_step = case_step.strip().split("\n")
    case_expected_result = case_expected_result.strip().split("\n")
    # 关键字
    case_keyword = ''
    # 优先级
    case_priority = gen_case_priority(testcase_dict['importance'])
    # 测试范围
    case_type = gen_case_type(testcase_dict['execution_type'])
    # 是否自动化
    case_apply_phase = gen_case_apply_phase(testcase_dict['summary'])
    # 需求单号
    # requirement_about_case = gen_case_root(testcase_dict['root'])
    # fileheader = ["所属模块", "用例标题", "前置条件", "步骤", "预期", "关键词", "优先级", "用例类型", "适用阶段"]
    # fileheader = ["用例名称", "需求号", "测试范围", "模块名称", "优先级", "为核心用例", "已自动化", "为回归用例",
    #              "前提条件", "操作步骤", "期望结果", "执行结果"]
    # row = [case_module, case_title, case_precontion, case_step, case_expected_result, case_keyword, case_priority,
    #        case_type, case_apply_phase]
    rows = []
    for i in range(0, len(case_step)):
        row1 = [case_title, requirement_about_case, case_type, case_module, case_priority, "否", case_apply_phase, "否",
                case_precontion, case_step[i], case_expected_result[i], "未执行"]
        logging.info("===========" * 40, row1)
        rows.append(row1)

    return rows


def gen_case_module(module_name):
    if module_name:
        module_name = module_name.replace('（', '(')
        module_name = module_name.replace('）', ')')
    else:
        module_name = '/'
    return module_name


def gen_case_step_and_expected_result(steps):
    case_step = ''
    case_expected_result = ''

    for step_dict in steps:
        case_step += str(step_dict['step_number']) + '. ' + step_dict['actions'].replace('\n', '').strip() + '\n'
        case_expected_result += str(step_dict['step_number']) + '. ' + \
                                step_dict['expectedresults'].replace('\n', '').strip() + '\n' \
            if step_dict.get('expectedresults', '') else ''

    return case_step, case_expected_result


def gen_case_priority(priority):
    mapping = {1: '紧急', 2: '一般', 3: '低'}
    if priority in mapping.keys():
        return mapping[priority]
    else:
        return '一般'


def gen_case_type(case_type):
    if case_type == '无':
        return "集成测试"
    else:
        return case_type


def gen_case_apply_phase(case_apply_phase):
    if case_apply_phase == '无':
        return "未自动化"
    else:
        return case_apply_phase


def gen_case_root(requirement_about_case):
    return requirement_about_case


def csv_to_xls(csv_path, xls_path):
    with open(csv_path, 'r', encoding='gb18030', errors='ignore') as f:
        data = f.read()
    data_file = StringIO(data)
    print(data_file)
    csv_reader = csv.reader(data_file)
    list_csv = []
    for row in csv_reader:
        list_csv.append(row)
    df_csv = pd.DataFrame(list_csv).applymap(str)
    '''
    这部分是不将csv装换为xls，而是过滤后再写入csv文件
    df_csv = df_csv[(df_csv[4] == '') | (df_csv[4] == 'name')]      # 过滤出第四列包含空值和name的数据
    df_csv.to_csv(csv_path, index=0, header=0, encoding='gb18030')  # 写入csv文件中
    '''
    writer = pd.ExcelWriter(xls_path)
    # 写入Excel
    df_csv.to_excel(
        excel_writer=writer,
        index=False,
        header=False
    )

    writer.save()
    # 删除csv文件
    os.remove(csv_path)

    # 合并表格相同项
    merge_nums = []
    list_keys = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I'}

    wb = load_workbook(xls_path)
    ws = wb.get_sheet_by_name('Sheet1')

    # 获取第一列数据
    type_list = []
    i = 2
    while True:
        r = ws.cell(i, 1).value
        if r:
            type_list.append(r)
        else:
            break
        i += 1

    # 判断合并单元格的始末位置
    s = 0
    e = 0
    flag = type_list[0]
    for i in range(len(type_list)):
        if type_list[i] != flag:
            flag = type_list[i]
            e = i - 1
            if e >= s:
                merge_nums.append([s + 2,e + 2])
                ws.merge_cells(list_keys[1] + str(s + 2) + ":" + list_keys[1] + str(e + 2))
                s = e + 1
        if i == len(type_list) - 1:
            e = i
            merge_nums.append([s + 2,e + 2])
            ws.merge_cells(list_keys[1] + str(s + 2) + ":" + list_keys[1] + str(e + 2))

    wb.save(xls_path)

    for j in range(2, 10):
        for i in merge_nums:
            ws.merge_cells(list_keys[j] + str(i[0]) + ":" + list_keys[j] + str(i[1]))

    wb.save(xls_path)

    return xls_path


if __name__ == '__main__':
    xmind_file = '../docs/zentao_testcase_template.xmind'
    yaxin_xls_file = xmind_to_yaxin_xls_file(xmind_file)
    print('Conver the xmind file to a yaxin xls file succssfully: %s', yaxin_xls_file)
